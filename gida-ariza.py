import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import os
import threading
from copy import copy # Hücre stillerini kopyalamak için gerekli

class AdvancedExcelAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gelişmiş Excel Keyword Analiz Aracı v6 - Stil Kopyalama")
        self.root.geometry("620x550")
        self.root.resizable(False, False)

        self.style = ttk.Style(self.root)
        self.style.theme_use("clam")

        self.excel_path = tk.StringVar()
        self.keywords_path = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        # --- Bu bölümde herhangi bir değişiklik yok ---
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        excel_frame = ttk.LabelFrame(main_frame, text="1. Adım: Kaynak Excel Dosyasını Seçin", padding="10")
        excel_frame.pack(fill=tk.X, pady=5)
        excel_entry = ttk.Entry(excel_frame, textvariable=self.excel_path, state="readonly", width=60)
        excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        excel_button = ttk.Button(excel_frame, text="Gözat...", command=self.select_excel_file)
        excel_button.pack(side=tk.LEFT)

        sheet_frame = ttk.LabelFrame(main_frame, text="2. Adım: İşlem Yapılacak Sayfayı Seçin", padding="10")
        sheet_frame.pack(fill=tk.X, pady=5)
        self.sheet_selector = ttk.Combobox(sheet_frame, state="disabled", width=65)
        self.sheet_selector.pack(fill=tk.X, expand=True)

        keywords_frame = ttk.LabelFrame(main_frame, text="3. Adım: Keyword Listesi (TXT) Seçin", padding="10")
        keywords_frame.pack(fill=tk.X, pady=5)
        keywords_entry = ttk.Entry(keywords_frame, textvariable=self.keywords_path, state="readonly", width=60)
        keywords_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        keywords_button = ttk.Button(keywords_frame, text="Gözat...", command=self.select_keywords_file)
        keywords_button.pack(side=tk.LEFT)
        
        settings_frame = ttk.LabelFrame(main_frame, text="4. Adım: Analiz Ayarları", padding="10")
        settings_frame.pack(fill=tk.X, pady=5)

        ttk.Label(settings_frame, text="Analiz Başlangıç Satırı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.start_row_entry = ttk.Entry(settings_frame, width=10)
        self.start_row_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        self.start_row_entry.insert(0, "4")

        ttk.Label(settings_frame, text="Aranacak Kolon Harfleri (örn: F):").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.source_columns_entry = ttk.Entry(settings_frame, width=20)
        self.source_columns_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        self.source_columns_entry.insert(0, "F")

        ttk.Label(settings_frame, text="Yeni Kolonun Konumu (Harf):").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        self.target_column_entry = ttk.Entry(settings_frame, width=10)
        self.target_column_entry.grid(row=2, column=1, padx=5, pady=5, sticky=tk.W)
        self.target_column_entry.insert(0, "C")
        ttk.Label(settings_frame, text="(Stilini sağındaki kolondan alır)").grid(row=2, column=2, padx=5, pady=5, sticky=tk.W)

        self.analyze_button = ttk.Button(main_frame, text="Analizi Başlat", command=self.start_analysis_thread, style="Accent.TButton")
        self.analyze_button.pack(pady=20, ipady=5, fill=tk.X)
        self.style.configure("Accent.TButton", font=("Helvetica", 10, "bold"))

        self.status_label = ttk.Label(main_frame, text="Durum: Lütfen bir Excel dosyası seçin.", anchor=tk.W)
        self.status_label.pack(fill=tk.X, pady=5, side=tk.BOTTOM)

    def select_excel_file(self):
        # --- Bu bölümde herhangi bir değişiklik yok ---
        path = filedialog.askopenfilename(title="Excel Dosyası Seçin", filetypes=(("Excel Dosyaları", "*.xlsx"),))
        if not path:
            return
        self.excel_path.set(path)
        self.status_label.config(text="Excel dosyası seçildi. Sayfalar okunuyor...")
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            self.sheet_selector['values'] = sheet_names
            self.sheet_selector.config(state="readonly")
            if sheet_names:
                self.sheet_selector.current(0)
            self.status_label.config(text="Excel dosyası ve sayfalar yüklendi. Lütfen diğer adımları tamamlayın.")
        except Exception as e:
            messagebox.showerror("Dosya Okuma Hatası", f"Excel sayfaları okunurken bir hata oluştu:\n{e}")
            self.status_label.config(text="Hata: Geçerli bir Excel dosyası seçilemedi.")
            self.sheet_selector.config(state="disabled")
            self.sheet_selector['values'] = []

    def select_keywords_file(self):
        # --- Bu bölümde herhangi bir değişiklik yok ---
        path = filedialog.askopenfilename(filetypes=(("Metin Dosyaları", "*.txt"),))
        if path:
            self.keywords_path.set(path)

    def start_analysis_thread(self):
        # --- Bu bölümde herhangi bir değişiklik yok ---
        self.analyze_button.config(state="disabled")
        self.status_label.config(text="Durum: Analiz başlatılıyor...")
        analysis_thread = threading.Thread(target=self.run_analysis, daemon=True)
        analysis_thread.start()

    def run_analysis(self):
        try:
            # --- GİRDİLERİ AL VE DOĞRULA ---
            excel_file = self.excel_path.get()
            selected_sheet_name = self.sheet_selector.get()
            keywords_file = self.keywords_path.get()
            start_row_str = self.start_row_entry.get()
            source_cols_str = self.source_columns_entry.get()
            target_col_str = self.target_column_entry.get().upper()

            if not all([excel_file, selected_sheet_name, keywords_file, start_row_str, source_cols_str, target_col_str]):
                raise ValueError("Lütfen tüm adımları eksiksiz doldurun!")
            if not start_row_str.isdigit() or int(start_row_str) < 2:
                raise ValueError("Başlangıç satırı en az '2' olmalıdır.")
            
            start_row = int(start_row_str)
            source_cols = [col.strip().upper() for col in source_cols_str.split(',')]
            target_col_index = column_index_from_string(target_col_str)

            # --- DOSYALARI YÜKLE ---
            self.status_label.config(text="Durum: Excel dosyası hafızaya yükleniyor...")
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb[selected_sheet_name]

            with open(keywords_file, 'r', encoding='utf-8') as f:
                keywords = [line.strip().lower() for line in f if line.strip()]

            # 1. ADIM: Mevcut birleştirilmiş hücreleri hafızaya al ve sonra kaldır
            self.status_label.config(text="Durum: Birleştirilmiş hücreler okunuyor...")
            original_merged_cells = [m.coord for m in sheet.merged_cells.ranges]
            for merged_range in list(sheet.merged_cells):
                 sheet.unmerge_cells(str(merged_range))

            # 2. ADIM: Hücreleri manuel olarak SAĞDAN SOLA doğru kaydır
            self.status_label.config(text="Durum: Yeni kolon için yer açılıyor...")
            max_col = sheet.max_column
            for col_idx in range(max_col, target_col_index - 1, -1):
                for row_idx in range(1, sheet.max_row + 1):
                    source_cell = sheet.cell(row=row_idx, column=col_idx)
                    if source_cell.value is not None or source_cell.has_style:
                        target_cell = sheet.cell(row=row_idx, column=col_idx + 1)
                        target_cell.value = source_cell.value
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.number_format = source_cell.number_format
                            target_cell.alignment = copy(source_cell.alignment)
                        source_cell.value = None
                        source_cell.style = 'Normal'

            # ----------------------------------------------------------------------------------
            # YENİ ADIM: Yeni Kolonu Sağındaki Gibi Stillendir
            # ----------------------------------------------------------------------------------
            self.status_label.config(text="Durum: Yeni kolon stillendiriliyor...")
            
            # 3.1: Sütun Genişliğini Kopyala
            source_col_letter = get_column_letter(target_col_index + 1)
            target_col_letter = get_column_letter(target_col_index)
            if source_col_letter in sheet.column_dimensions:
                sheet.column_dimensions[target_col_letter].width = sheet.column_dimensions[source_col_letter].width
            
            # 3.2: Hücre Stillerini (Kenarlık, Renk vb.) Satır Satır Kopyala
            for row_idx in range(1, sheet.max_row + 1):
                source_style_cell = sheet.cell(row=row_idx, column=target_col_index + 1)
                target_style_cell = sheet.cell(row=row_idx, column=target_col_index)
                
                if source_style_cell.has_style:
                    target_style_cell.font = copy(source_style_cell.font)
                    target_style_cell.border = copy(source_style_cell.border)
                    target_style_cell.fill = copy(source_style_cell.fill)
                    target_style_cell.number_format = source_style_cell.number_format
                    target_style_cell.alignment = copy(source_style_cell.alignment)
            # ----------------------------------------------------------------------------------

            # 4. ADIM: Analizi yap ve stillendirilmiş yeni kolona veriyi yaz
            self.status_label.config(text="Durum: Satırlar analiz ediliyor...")
            header_row = start_row - 1
            # Başlık hücresinin artık stili var, sadece üzerine yazıyoruz
            sheet.cell(row=header_row, column=target_col_index).value = "BULUNAN KELİMELER"

            for row_index in range(start_row, sheet.max_row + 1):
                found_keywords = set()
                for col_letter in source_cols:
                    col_idx_before_shift = column_index_from_string(col_letter)
                    col_idx_after_shift = col_idx_before_shift
                    if col_idx_before_shift >= target_col_index:
                        col_idx_after_shift += 1
                    
                    cell_value = sheet.cell(row=row_index, column=col_idx_after_shift).value
                    cell_value_str = str(cell_value).lower() if cell_value else ""

                    for keyword in keywords:
                        if keyword in cell_value_str:
                            found_keywords.add(keyword)
                
                if found_keywords:
                    result_str = ", ".join(sorted(list(found_keywords)))
                    sheet.cell(row=row_index, column=target_col_index).value = result_str

            # 5. ADIM: Hafızadaki birleştirilmiş hücreleri YENİ KOORDİNATLARLA tekrar uygula
            self.status_label.config(text="Durum: Birleştirilmiş hücreler düzeltiliyor...")
            for merged_coord in original_merged_cells:
                min_col_str, min_row_str, max_col_str, max_row_str = openpyxl.utils.cell.range_boundaries(merged_coord)
                
                new_min_col, new_max_col = min_col_str, max_col_str
                if min_col_str >= target_col_index: new_min_col += 1
                if max_col_str >= target_col_index: new_max_col += 1
                
                new_range = f"{get_column_letter(new_min_col)}{min_row_str}:{get_column_letter(new_max_col)}{max_row_str}"
                sheet.merge_cells(new_range)
            
            # --- SONUCU YENİ DOSYAYA KAYDET ---
            self.status_label.config(text="Durum: Sonuçlar yeni dosyaya kaydediliyor...")
            output_dir = os.path.dirname(excel_file)
            base_name = os.path.basename(excel_file)
            file_name, file_ext = os.path.splitext(base_name)
            output_path = os.path.join(output_dir, f"{file_name}_sonuc.xlsx")
            
            wb.save(output_path)
            
            self.status_label.config(text=f"Başarılı! Dosya '{os.path.basename(output_path)}' olarak kaydedildi.")
            messagebox.showinfo("İşlem Tamamlandı", f"Analiz başarıyla tamamlandı.\nFormatlama ve stiller korunarak sonuç dosyası oluşturuldu:\n{output_path}")

        except Exception as e:
            messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")
            self.status_label.config(text=f"Hata: {e}")
        finally:
            if 'analyze_button' in self.__dict__ and self.analyze_button.winfo_exists():
                self.analyze_button.config(state="normal")


if __name__ == "__main__":
    root = tk.Tk()
    app = AdvancedExcelAnalyzerApp(root)
    root.mainloop()